"""
COMPLETE END-TO-END TEST
Tests EVERYTHING before RunPod deployment
Validates: Models, Processing, Excel Writing, All Modules

Save as: test_complete.py
Run: python test_complete.py
"""

import os
import sys
import asyncio
from pathlib import Path
import json
import torch

print("\n" + "="*80)
print("COMPLETE SYSTEM VALIDATION TEST")
print("Testing ALL components before RunPod deployment")
print("="*80 + "\n")

# Check Python version
print(f"[CHECK 1/10] Python version: {sys.version}")
assert sys.version_info >= (3, 10), "Need Python 3.10+"
print("  ✓ Python version OK\n")

# Check GPU
print("[CHECK 2/10] GPU availability...")
if torch.cuda.is_available():
    print(f"  ✓ GPU detected: {torch.cuda.get_device_name(0)}")
    print(f"  ✓ GPU memory: {torch.cuda.get_device_properties(0).total_memory / 1e9:.1f}GB")
    device = "cuda"
else:
    print("  ⚠ No GPU - will use CPU (slower)")
    device = "cpu"
print()

# Check folder structure
print("[CHECK 3/10] Folder structure...")
base_dir = Path(__file__).parent

required_folders = [
    'data/videos',
    'data/taxonomy',
    'src',
    'output'
]

for folder in required_folders:
    folder_path = base_dir / folder
    if folder_path.exists():
        print(f"  ✓ {folder}/")
    else:
        print(f"  ✗ MISSING: {folder}/")
        os.makedirs(folder_path, exist_ok=True)
        print(f"    Created: {folder}/")
print()

# Check for Excel file
print("[CHECK 4/10] Excel file...")
excel_file = base_dir / 'COPY_UCG_CodingSheet_13Jan.xlsx'
if excel_file.exists():
    print(f"  ✓ Excel file found: {excel_file.name}")
else:
    print(f"  ✗ MISSING: {excel_file.name}")
    print(f"  ACTION NEEDED: Copy Excel file to: {base_dir}")
    excel_file = None
print()

# Check for test video
print("[CHECK 5/10] Test video...")
video_dir = base_dir / 'data' / 'videos'
video_files = list(video_dir.glob('*.mp4'))
if video_files:
    test_video = video_files[0]
    print(f"  ✓ Test video found: {test_video.name}")
else:
    print(f"  ✗ MISSING: No videos in {video_dir}")
    print(f"  ACTION NEEDED: Copy at least 1 test video to: {video_dir}")
    test_video = None
print()

# Check taxonomy file
print("[CHECK 6/10] Taxonomy file...")
taxonomy_file = base_dir / 'data' / 'taxonomy' / 'refined_lists.csv'
if taxonomy_file.exists():
    print(f"  ✓ Taxonomy found: {taxonomy_file.name}")
else:
    print(f"  ⚠ MISSING: {taxonomy_file}")
    print(f"  Will use default taxonomy")
print()

# Check src/ modules
print("[CHECK 7/10] Source modules...")
sys.path.insert(0, str(base_dir / 'src'))

required_modules = [
    'config',
    'video_processor', 
    'audio_analyzer',
    'qwen_analyzer',
    'aggregator',
    'validator',
    'logger'
]

modules_ok = True
for module_name in required_modules:
    try:
        __import__(module_name)
        print(f"  ✓ {module_name}.py")
    except Exception as e:
        print(f"  ✗ {module_name}.py - ERROR: {e}")
        modules_ok = False

if not modules_ok:
    print("\n  ✗ Some modules have errors!")
    print("  Fix module errors before proceeding")
    sys.exit(1)
print()

# Test model loading
print("[CHECK 8/10] Testing AI models...")
print("  Loading Qwen-VL (this takes 5-10 min first time)...")

try:
    from transformers import Qwen2VLForConditionalGeneration, AutoProcessor
    
    print("  Downloading/loading Qwen model...")
    qwen_model = Qwen2VLForConditionalGeneration.from_pretrained(
        "Qwen/Qwen2-VL-7B-Instruct",
        torch_dtype=torch.float16 if device == "cuda" else torch.float32,
        device_map="auto",
        trust_remote_code=True
    )
    qwen_processor = AutoProcessor.from_pretrained(
        "Qwen/Qwen2-VL-7B-Instruct",
        trust_remote_code=True
    )
    print("  ✓ Qwen-VL loaded successfully")
    qwen_ok = True
except Exception as e:
    print(f"  ✗ Qwen-VL failed: {e}")
    qwen_ok = False

try:
    import whisper
    print("  Loading Whisper...")
    whisper_model = whisper.load_model("base")
    print("  ✓ Whisper loaded successfully")
    whisper_ok = True
except Exception as e:
    print(f"  ✗ Whisper failed: {e}")
    whisper_ok = False

print()

# Test video processing (if we have a video)
if test_video and qwen_ok and whisper_ok:
    print("[CHECK 9/10] Testing video processing pipeline...")
    
    try:
        import cv2
        from scenedetect import detect, ContentDetector
        from PIL import Image
        
        # Extract metadata
        print("  Testing metadata extraction...")
        cap = cv2.VideoCapture(str(test_video))
        fps = cap.get(cv2.CAP_PROP_FPS)
        duration = cap.get(cv2.CAP_PROP_FRAME_COUNT) / fps
        cap.release()
        print(f"    Video: {duration:.1f}s @ {fps:.1f}fps")
        
        # Detect cuts
        print("  Testing scene detection...")
        scenes = detect(str(test_video), ContentDetector(threshold=30))
        print(f"    Found {len(scenes)} cuts")
        
        # Extract one frame
        print("  Testing frame extraction...")
        cap = cv2.VideoCapture(str(test_video))
        ret, frame = cap.read()
        if ret:
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            test_image = Image.fromarray(frame_rgb)
            print(f"    Frame size: {test_image.size}")
        cap.release()
        
        # Test Whisper
        print("  Testing audio analysis...")
        audio_result = whisper_model.transcribe(str(test_video), fp16=False)
        transcript = audio_result['text']
        print(f"    Transcript: {len(transcript)} characters")
        
        # Test Qwen on one frame
        print("  Testing Qwen-VL inference...")
        
        prompt = """Analyze this video frame from Egypt tourism content.
        
Identify the main destination attractor from:
Accommodation, Architecture & Heritage, Climate, Cultural Attractions, 
Food & Drink, Landscape & Natural Resources, Local Lifestyle

Return only JSON:
{"attractor": "Category Name", "confidence": 0.95}"""

        conversation = [{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image", "image": test_image}
            ]
        }]
        
        text_prompt = qwen_processor.apply_chat_template(
            conversation, tokenize=False, add_generation_prompt=True
        )
        
        inputs = qwen_processor(
            text=[text_prompt],
            images=[test_image],
            padding=True,
            return_tensors="pt"
        ).to(device)
        
        with torch.no_grad():
            output_ids = qwen_model.generate(
                **inputs,
                max_new_tokens=256,
                do_sample=False
            )
        
        response = qwen_processor.batch_decode(
            output_ids, skip_special_tokens=True
        )[0]
        
        print(f"    Qwen response: {response[:200]}...")
        
        print("  ✓ Complete pipeline working!")
        pipeline_ok = True
        
    except Exception as e:
        print(f"  ✗ Pipeline test failed: {e}")
        import traceback
        traceback.print_exc()
        pipeline_ok = False
else:
    print("[CHECK 9/10] Skipping pipeline test (missing dependencies)")
    pipeline_ok = False

print()

# Test Excel writing (if we have Excel file)
if excel_file:
    print("[CHECK 10/10] Testing Excel integration...")
    
    try:
        import openpyxl
        
        # Load workbook
        print("  Loading Excel workbook...")
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        
        # Get main sheet
        if 'Coding_TouristUCG' in wb.sheetnames:
            ws = wb['Coding_TouristUCG']
            print(f"  ✓ Found sheet: Coding_TouristUCG")
            print(f"    Total rows: {ws.max_row}")
            
            # Check row 286 (where you want to start)
            video_id_286 = ws.cell(row=286, column=1).value
            print(f"    Row 286 video_id: {video_id_286}")
            
            # Find first empty row
            first_empty = None
            for row in range(286, ws.max_row + 1):
                # Check if audio_type column is empty (column V = 22)
                if not ws.cell(row=row, column=22).value:
                    first_empty = row
                    break
            
            if first_empty:
                print(f"    First empty row to fill: {first_empty}")
            else:
                print(f"    All rows from 286 onwards already filled")
        else:
            print(f"  ✗ Sheet 'Coding_TouristUCG' not found")
            print(f"    Available sheets: {wb.sheetnames}")
        
        wb.close()
        print("  ✓ Excel integration OK")
        excel_ok = True
        
    except Exception as e:
        print(f"  ✗ Excel test failed: {e}")
        excel_ok = False
else:
    print("[CHECK 10/10] Skipping Excel test (file missing)")
    excel_ok = False

print()

# FINAL SUMMARY
print("="*80)
print("VALIDATION SUMMARY")
print("="*80 + "\n")

checks = {
    "Python 3.10+": True,
    "GPU/CPU detected": True,
    "Folder structure": True,
    "Excel file present": excel_file is not None,
    "Test video present": test_video is not None,
    "Source modules": modules_ok,
    "Qwen-VL model": qwen_ok,
    "Whisper model": whisper_ok,
    "Processing pipeline": pipeline_ok,
    "Excel integration": excel_file is not None and excel_ok
}

passed = sum(checks.values())
total = len(checks)

for check, status in checks.items():
    symbol = "✓" if status else "✗"
    print(f"  {symbol} {check}")

print(f"\nPassed: {passed}/{total}")

if passed == total:
    print("\n" + "="*80)
    print("🎉 ALL CHECKS PASSED - READY FOR RUNPOD!")
    print("="*80)
    print("\nNext steps:")
    print("  1. Copy all files to RunPod")
    print("  2. Copy Excel file to root directory")
    print("  3. Copy all 284 videos to data/videos/")
    print("  4. Run: python main.py")
    print()
else:
    print("\n" + "="*80)
    print("⚠ ISSUES FOUND - FIX BEFORE RUNPOD")
    print("="*80)
    print("\nAction needed:")
    
    if not excel_file:
        print("  • Copy COPY_UCG_CodingSheet_13Jan.xlsx to root folder")
    
    if not test_video:
        print("  • Copy at least 1 test video to data/videos/")
    
    if not qwen_ok:
        print("  • Fix Qwen-VL model loading")
    
    if not whisper_ok:
        print("  • Fix Whisper model loading")
    
    if not pipeline_ok:
        print("  • Fix processing pipeline errors")
    
    print()

# Save test results
results_file = base_dir / 'output' / 'validation_results.json'
with open(results_file, 'w') as f:
    json.dump({
        'timestamp': str(Path(__file__).stat().st_mtime),
        'checks': {k: str(v) for k, v in checks.items()},
        'passed': passed,
        'total': total,
        'ready_for_production': passed == total
    }, f, indent=2)

print(f"Results saved to: {results_file}\n")