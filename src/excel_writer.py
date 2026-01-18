"""
FORMULA-SAFE EXCEL WRITER
Writes AI results to Excel without breaking formulas
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd
from typing import Dict, List
import re

class ExcelWriter:
    """
    Writes video analysis results to Excel workbook
    CRITICAL: Preserves all formula cells
    """
    
    def __init__(self, config):
        self.config = config
        self.workbook = None
        self.main_sheet = None
        self.taxonomy_sheet = None
        
        # Column mappings (from your screenshots)
        self.column_map = self._define_column_mappings()
        
    def load_workbook(self):
        """Load existing Excel workbook"""
        self.workbook = openpyxl.load_workbook(
            self.config.excel_path,
            data_only=False  # Keep formulas
        )
        
        # Get sheets
        self.main_sheet = self.workbook['Coding_TouristUCG']
        self.taxonomy_sheet = self.workbook['Refined_Lists']
        
        print(f"✓ Loaded Excel: {self.config.excel_path}")
        print(f"  Main sheet rows: {self.main_sheet.max_row}")
        
    def load_taxonomy(self) -> Dict:
        """Load taxonomy from Refined_Lists sheet"""
        taxonomy = {
            'attractors': {},
            'visual_tags': {},
            'audio_tags': {},
            'text_tags': {}
        }
        
        # Parse taxonomy sheet
        # This would read your predefined lists
        # For now, returning structure
        
        return taxonomy
    
    def get_video_list(self) -> List[Dict]:
        """
        Get list of videos to process from Excel
        Returns list of {video_id, video_file_name}
        """
        videos = []
        
        # Start from row 2 (after headers)
        for row in range(2, self.main_sheet.max_row + 1):
            video_id_cell = self.main_sheet.cell(
                row=row, 
                column=self.column_map['video_id']
            )
            
            filename_cell = self.main_sheet.cell(
                row=row,
                column=self.column_map['video_file_name']
            )
            
            video_id = video_id_cell.value
            filename = filename_cell.value
            
            if video_id and filename:
                videos.append({
                    'video_id': video_id,
                    'video_file_name': filename,
                    'row_number': row
                })
        
        return videos
    
    def write_video_result(self, result: Dict):
        """
        Write analysis result for one video to Excel
        CRITICAL: Only write to non-formula cells
        """
        
        # Find row by video_id
        video_id = result['video_id']
        row_num = self._find_video_row(video_id)
        
        if row_num is None:
            print(f"Warning: Video {video_id} not found in Excel")
            return
        
        # Technical metrics
        self._write_cell(row_num, 'audio_type', result.get('audio_type', ''))
        self._write_cell(row_num, 'text_overlay_count', result.get('text_overlay_count', 0))
        self._write_cell(row_num, 'total_cuts', result.get('total_cuts', 0))
        
        # Attractor seconds (19 columns)
        attractor_seconds = result.get('attractor_seconds', {})
        for attractor_name, seconds in attractor_seconds.items():
            col_key = f"{attractor_name.lower().replace(' ', '_').replace('&', 'and')}_seconds"
            if col_key in self.column_map:
                self._write_cell(row_num, col_key, round(seconds, 2))
        
        # Attractor subcategories and places
        subcategories = result.get('subcategories', {})
        places = result.get('places', {})
        
        for attractor, subcat in subcategories.items():
            col_key = f"{attractor.lower().replace(' ', '_')}_subcategory"
            if col_key in self.column_map:
                self._write_cell(row_num, col_key, subcat)
        
        for attractor, place in places.items():
            col_key = f"{attractor.lower().replace(' ', '_')}_place"
            if col_key in self.column_map:
                self._write_cell(row_num, col_key, place)
        
        # Narrative frames - Visual modality
        visual = result.get('narrative_visual', {})
        visual_tags = result.get('tags_visual', {})
        
        for frame in ['AE', 'LI', 'IN', 'CR']:
            # Seconds
            self._write_cell(
                row_num, 
                f'{frame}_visual_seconds',
                round(visual.get(frame, 0), 2)
            )
            # Tags
            self._write_cell(
                row_num,
                f'{frame}_visual_tag',
                visual_tags.get(frame, '')
            )
        
        # Narrative frames - Audio modality
        audio = result.get('narrative_audio', {})
        audio_tags = result.get('tags_audio', {})
        
        for frame in ['AE', 'LI', 'IN', 'CR']:
            self._write_cell(
                row_num,
                f'{frame}_audio_seconds',
                round(audio.get(frame, 0), 2)
            )
            self._write_cell(
                row_num,
                f'{frame}_audio_tag',
                audio_tags.get(frame, '')
            )
        
        # Narrative frames - Text modality
        text = result.get('narrative_text', {})
        text_tags = result.get('tags_text', {})
        
        for frame in ['AE', 'LI', 'IN', 'CR']:
            self._write_cell(
                row_num,
                f'{frame}_text_seconds',
                round(text.get(frame, 0), 2)
            )
            self._write_cell(
                row_num,
                f'{frame}_text_tag',
                text_tags.get(frame, '')
            )
        
        # Save after each video (incremental)
        self.workbook.save(self.config.excel_output_path)
    
    def _write_cell(self, row: int, column_key: str, value):
        """
        Write to a cell ONLY if it's not a formula
        """
        if column_key not in self.column_map:
            return
        
        col_num = self.column_map[column_key]
        cell = self.main_sheet.cell(row=row, column=col_num)
        
        # Check if cell contains formula
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  WARNING: Skipping formula cell {get_column_letter(col_num)}{row}")
            return
        
        # Safe to write
        cell.value = value
    
    def _find_video_row(self, video_id: str) -> int:
        """Find row number for a video_id"""
        video_id_col = self.column_map['video_id']
        
        for row in range(2, self.main_sheet.max_row + 1):
            cell_value = self.main_sheet.cell(row=row, column=video_id_col).value
            if cell_value == video_id:
                return row
        
        return None
    
    def save_workbook(self):
        """Save the workbook"""
        self.workbook.save(self.config.excel_output_path)
        print(f"✓ Excel saved: {self.config.excel_output_path}")
    
    def _define_column_mappings(self) -> Dict:
        """
        Define exact column positions from Excel
        Based on your screenshots
        """
        return {
            # Metadata (Column A-J approximate)
            'video_id': 1,  # Column A
            'video_file_name': 2,  # Column B
            'coding_date': 4,  # Column D
            'video_length_sec': 9,  # Column I
            
            # Technical (columns around V-X based on screenshot)
            'audio_type': 22,  # Column V
            'text_overlay_count': 23,  # Column W
            'total_cuts': 24,  # Column X
            
            # Attractor Seconds (19 columns starting around Y)
            # Based on your screenshot showing these columns
            'accommodation_seconds': 25,
            'architecture_and_heritage_seconds': 26,
            'climate_seconds': 27,
            'cultural_attractions_seconds': 28,
            'events_fairs_and_festivals_seconds': 29,
            'food_and_drink_seconds': 30,
            'infrastructure_and_transportation_seconds': 31,
            'landscape_and_natural_resources_seconds': 32,
            'leisure_attractions_seconds': 33,
            'local_culture_and_history_seconds': 34,
            'local_lifestyle_seconds': 35,
            'nightlife_seconds': 36,
            'political_and_economic_factors_seconds': 37,
            'safety_seconds': 38,
            'service_seconds': 39,
            'shopping_seconds': 40,
            'sports_seconds': 41,
            'tourism_products_and_packages_seconds': 42,
            'wellness_seconds': 43,
            
            # Percentages are FORMULAS - don't touch (columns 44-62)
            
            # Subcategories & Places (columns 63-100 based on pattern)
            # These would be defined for each attractor
            
            # Narrative Frame columns (starting around column 105+ based on screenshots)
            # AE frame
            'AE_visual_tag': 105,
            'AE_visual_seconds': 106,
            'AE_text_tag': 108,
            'AE_text_seconds': 109,
            'AE_audio_tag': 111,
            'AE_audio_seconds': 112,
            
            # LI frame
            'LI_visual_tag': 115,
            'LI_visual_seconds': 116,
            'LI_text_tag': 118,
            'LI_text_seconds': 119,
            'LI_audio_tag': 121,
            'LI_audio_seconds': 122,
            
            # IN frame
            'IN_visual_tag': 125,
            'IN_visual_seconds': 126,
            'IN_text_tag': 128,
            'IN_text_seconds': 129,
            'IN_audio_tag': 131,
            'IN_audio_seconds': 132,
            
            # CR frame
            'CR_visual_tag': 135,
            'CR_visual_seconds': 136,
            'CR_audio_tag': 138,
            'CR_audio_seconds': 139,
            'CR_text_tag': 141,
            'CR_text_seconds': 142
        }